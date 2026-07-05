import os
import re
from database import get_db_connection, get_cursor
from io import BytesIO
from datetime import datetime
from typing import List, Optional

import pandas as pd
import numpy as np
from fastapi import FastAPI, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from rapidfuzz import process, fuzz

from sklearn.feature_extraction.text import TfidfVectorizer

# Import ranking logic
from ranking import calculate_rankings, get_work_theme
from auth import router as auth_router

# Initialize FastAPI App
app = FastAPI(
    title="People's Priorities AI API",
    description="Backend API serving the Thematic Prioritization framework for Hackathons.",
    version="1.0.0"
)

# CORS middleware configuration (Hackathon-safe, not production-ready)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Include Authentication router
app.include_router(auth_router)

@app.get("/")
def read_root():
    return {"status": "healthy", "service": "peoples-priorities-api"}

@app.get("/admin/seed")
def seed_database_route():
    try:
        from database import DATABASE_URL
        import psycopg2
        from psycopg2.extras import RealDictCursor
        
        # Force Postgres connection directly to inspect connection issues on Vercel
        try:
            conn = psycopg2.connect(DATABASE_URL, connect_timeout=5)
            conn.autocommit = True
            cursor = conn.cursor(cursor_factory=RealDictCursor)
        except Exception as pg_err:
            import traceback
            import re
            tb = traceback.format_exc()
            masked_url = "None"
            if DATABASE_URL:
                masked_url = re.sub(r':([^@]+)@', ':****@', DATABASE_URL)
            return {
                "status": "error", 
                "message": f"Postgres connection failed: {str(pg_err)}",
                "database_url_seen": masked_url,
                "traceback": tb
            }
        
        # 1. Create tables if they do not exist
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS villages (
            village_id TEXT PRIMARY KEY,
            village_name TEXT NOT NULL,
            subdistrict TEXT NOT NULL,
            population INTEGER NOT NULL,
            literacy_rate_pct REAL NOT NULL,
            sc_st_pct REAL NOT NULL,
            distance_to_town_km REAL NOT NULL,
            has_primary_school TEXT NOT NULL,
            has_secondary_school TEXT NOT NULL,
            has_health_subcentre TEXT NOT NULL,
            has_tap_water TEXT NOT NULL,
            has_paved_road TEXT NOT NULL
        );
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS schools (
            udise_code TEXT PRIMARY KEY,
            village_id TEXT NOT NULL,
            school_category TEXT NOT NULL,
            total_enrollment INTEGER NOT NULL,
            building_capacity INTEGER NOT NULL,
            enrollment_capacity_ratio REAL NOT NULL,
            pupil_teacher_ratio INTEGER NOT NULL,
            functional_toilets INTEGER NOT NULL,
            FOREIGN KEY (village_id) REFERENCES villages (village_id)
        );
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS existing_works (
            work_id TEXT PRIMARY KEY,
            village_id TEXT NOT NULL,
            source TEXT NOT NULL,
            work_type TEXT NOT NULL,
            amount_lakh_inr REAL NOT NULL,
            status TEXT NOT NULL,
            year INTEGER NOT NULL,
            FOREIGN KEY (village_id) REFERENCES villages (village_id)
        );
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS submissions (
            submission_id TEXT PRIMARY KEY,
            village_id TEXT NOT NULL,
            raw_text TEXT NOT NULL,
            channel TEXT NOT NULL,
            language_detected TEXT NOT NULL,
            submitted_on TEXT NOT NULL,
            status TEXT NOT NULL,
            FOREIGN KEY (village_id) REFERENCES villages (village_id)
        );
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS themes (
            theme_id SERIAL PRIMARY KEY,
            theme_label TEXT NOT NULL UNIQUE,
            keyword_summary TEXT
        );
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS submission_themes (
            submission_id TEXT NOT NULL,
            theme_id INTEGER NOT NULL,
            confidence_score REAL NOT NULL,
            PRIMARY KEY (submission_id, theme_id),
            FOREIGN KEY (submission_id) REFERENCES submissions (submission_id),
            FOREIGN KEY (theme_id) REFERENCES themes (theme_id)
        );
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        
        conn.commit()
        
        # 2. Seed data
        # Check if villages table has data already to prevent duplicates
        cursor.execute("SELECT COUNT(*) as count FROM villages;")
        count_res = cursor.fetchone()
        
        # Support dict format (psycopg2 RealDictCursor) or tuple (standard cursor)
        count = count_res['count'] if isinstance(count_res, dict) else count_res[0]
        
        if count == 0:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            data_dir = os.path.join(base_dir, "data")
            
            from sqlalchemy import create_engine
            engine = create_engine(DATABASE_URL)
            
            csv_to_table = {
                "villages.csv": "villages",
                "schools.csv": "schools",
                "existing_works.csv": "existing_works",
                "submissions.csv": "submissions"
            }
            
            for csv_file, table_name in csv_to_table.items():
                csv_path = os.path.join(data_dir, csv_file)
                if os.path.exists(csv_path):
                    df = pd.read_csv(csv_path)
                    df.to_sql(table_name, engine, if_exists="append", index=False)
            
            conn.close()
            return {"status": "success", "message": "Database tables created and seeded successfully!"}
        else:
            conn.close()
            return {"status": "success", "message": f"Database tables already exist. Row count in 'villages' is {count}."}
            
    except Exception as e:
        return {"status": "error", "message": str(e)}

# Hardcoded dict of themes and keywords for classification fallback and label matching
themes_keywords = {
    "School Infrastructure": ["school", "classroom", "toilet", "bench", "desk", "playground", "blackboard", "building", "library", "anganwadi", "kitchen", "teachers", "computer", "lab", "pathshala", "vidyalaya", "shauchalay", "chat", "स्कूल", "शिक्षिका", "शिक्षक", "कक्षा", "शौचालय", "डेस्क", "बेंच", "पुस्तकालय", "आंगनवाड़ी", "ब्लैकबोर्ड"],
    "Water Supply": ["water", "drinking", "pump", "handpump", "borewell", "pipeline", "tap", "tank", "leak", "fluoride", "well", "potable", "filtration", "paani", "jal", "nal", "tanki", "kuan", "kuon", "पानी", "नल", "पाइपलाइन", "फ्लोराइड", "बोरवेल", "टंकी", "हैंडपंप", "कुआं"],
    "Road Connectivity": ["road", "pothole", "highway", "bridge", "paved", "streetlight", "bus", "cremation", "drainage", "PMGSY", "mud", "monsoon", "sadak", "rasta", "kichad", "gaddha", "gaddhe", "pul", "naali", "सड़क", "रास्ता", "कीचड़", "गड्ढे", "पुल", "नाली", "लाइट", "खड़ंजा", "शमशान"],
    "Healthcare": ["health", "subcentre", "clinic", "doctor", "nurse", "midwife", "ambulance", "medicine", "vaccination", "treatment", "hospital", "ANM", "PHC", "dawa", "dawaiya", "tika", "tikakaran", "bimari", "doctor", "स्वास्थ्य", "अस्पताल", "इलाज", "डॉक्टर", "नर्स", "दवाई", "दवाइयां", "टीकाकरण", "टीका", "एंबुलेंस", "बीमारी"],
    "Employment/Skills": ["mgnrega", "nrega", "job", "wage", "work", "payment", "skills", "training", "tailoring", "computer", "SHG", "poultry", "dairy", "employment", "migration", "rozgar", "nokri", "kam", "kaam", "pesa", "paisa", "bhatta", "मनरेगा", "रोजगार", "काम", "नौकरी", "मजदूरी", "भुगतान", "प्रशिक्षण", "ट्रेनिंग", "मस्टर"],
    "Electricity": ["electricity", "power", "transformer", "voltage", "streetlight", "bill", "connection", "wire", "poles", "solar", "grid", "light", "bijli", "line", "fluctuation", "meter", "बिजली", "ट्रांसफार्मर", "वोल्टेज", "बिल", "तार", "खंभे", "सोलर", "पावर", "लाइट", "मीटर"],
    "Irrigation": ["irrigation", "canal", "dam", "borewell", "pump", "drip", "pond", "reservoir", "desilting", "groundwater", "harvest", "recharge", "sinchai", "nahar", "khet", "talab", "paani", "fasal", "sookh", "सिंचाई", "नहर", "डेम", "चेक डैम", "फसल", "सूख", "तालाब", "वॉटर", "रिचार्ज", "नलकूप"]
}

# Stopwords for cleaning up TF-IDF keyword extraction
stop_words = [
    'hai', 'ki', 'ko', 'me', 'se', 'ke', 'ka', 'humare', 'gaon', 'ho', 'par', 'itne', 
    'hote', 'sath', 'aur', 'bhi', 'tha', 'thi', 'he', 'ye', 'pe', 'huare', 'bohot', 
    'lekin', 'kar', 'raha', 'rahe', 'sahi', 'sath', 'is', 'the', 'to', 'for', 'of', 
    'in', 'and', 'we', 'our', 'has', 'have', 'are', 'a', 'an', 'at', 'on', 'with', 
    'from', 'by', 'but', 'not', 'ke', 'mein', 'hai', 'ki', 'ko', 'se', 'ka', 'aur', 
    'bhi', 'tha', 'thi', 'hain', 'ho', 'कर', 'रहे', 'रहा', 'लिए', 'पर', 'नहीं', 
    'सकते', 'सकता', 'गया', 'गई', 'वार्ड', 'तरह', 'बहुत', 'कम', 'करके', 'करने', 'लोग'
]

# Database Path Resolution
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "data", "peoples_priorities.db")

# -------------------------------------------------------------------------
# Helper Functions
# -------------------------------------------------------------------------

def clean_and_tokenize(text):
    text = text.lower()
    words = re.findall(r'\b\w+\b', text)
    return words

def get_or_create_theme(conn, label, keywords):
    cursor = get_cursor(conn)
    cursor.execute("SELECT theme_id FROM themes WHERE theme_label = %s", (label,))
    row = cursor.fetchone()
    if row:
        return row['theme_id']
    cursor.execute("INSERT INTO themes (theme_label, keyword_summary) VALUES (%s, %s) RETURNING theme_id", (label, keywords))
    conn.commit()
    return cursor.fetchone()['theme_id']

# -------------------------------------------------------------------------
# Pydantic Schemas
# -------------------------------------------------------------------------
class SubmissionCreate(BaseModel):
    raw_text: str = Field(..., min_length=10, description="The content of the complaint/priority.")
    village_name: str = Field(..., description="Village name to match against metadata.")
    channel: str = Field(..., description="Channel of entry e.g. web_form, whatsapp, sms.")
    language_detected: str = Field(..., description="Language of text e.g. English, Hindi, Hinglish.")
    media_url: Optional[str] = Field(None, description="Optional attachment link.")

class ProcessSubmissionsRequest(BaseModel):
    submission_ids: Optional[List[str]] = Field(None, description="List of submission IDs to process. Empty = process all unprocessed submissions.")

class BudgetSimulationRequest(BaseModel):
    budget_lakh_inr: float = Field(..., description="Total available budget in lakhs")
    theme_filter: Optional[str] = Field(None, description="Optional theme_id or theme_label to restrict to one theme only")
    village_filter: Optional[str] = Field(None, description="Optional village_id or village_name to restrict to one village only")

class BudgetSimulationExportRequest(BudgetSimulationRequest):
    format: str = Field(..., pattern="^(pdf|xlsx)$", description="Format to export: pdf or xlsx")

# -------------------------------------------------------------------------
# Endpoints
# -------------------------------------------------------------------------

@app.get("/health", status_code=status.HTTP_200_OK)
def health_check():
    return {"status": "ok", "timestamp": datetime.utcnow().isoformat()}

@app.post("/submissions", status_code=status.HTTP_201_CREATED)
def create_submission(payload: SubmissionCreate):
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        # Load villages for fuzzy matching
        cursor.execute("SELECT village_id, village_name FROM villages")
        village_rows = cursor.fetchall()
        
        choices = {row['village_name']: row['village_id'] for row in village_rows}
        
        # Fuzzy Match
        best_match = process.extractOne(payload.village_name, choices.keys(), scorer=fuzz.WRatio)
        if not best_match:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, 
                detail=f"Village '{payload.village_name}' could not be matched."
            )
            
        matched_name, score, _ = best_match
        if score < 60:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, 
                detail=f"Village '{payload.village_name}' matches '{matched_name}' with low confidence score ({score:.1f})."
            )
            
        village_id = choices[matched_name]
        
        # Generate new sequential submission_id (e.g. SUB088)
        cursor.execute("SELECT submission_id FROM submissions")
        sub_rows = cursor.fetchall()
        
        max_num = 0
        for row in sub_rows:
            val = row['submission_id']
            if val.startswith("SUB"):
                try:
                    num = int(val[3:])
                    if num > max_num:
                        max_num = num
                except ValueError:
                    pass
        new_id = f"SUB{max_num + 1:03d}"
        
        # Insert submission
        cursor.execute("""
            INSERT INTO submissions (submission_id, village_id, raw_text, channel, language_detected, submitted_on, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            new_id,
            village_id,
            payload.raw_text,
            payload.channel,
            payload.language_detected,
            datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
            "received"
        ))
        conn.commit()
        
        return {"submission_id": new_id, "status": "received"}
        
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {e}")
    finally:
        conn.close()

@app.post("/submissions/process")
def process_submissions(payload: ProcessSubmissionsRequest):
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        # Determine target IDs to process
        if payload.submission_ids:
            cursor.execute(
                f"SELECT submission_id, raw_text FROM submissions WHERE submission_id IN ({','.join(['?']*len(payload.submission_ids))})",
                payload.submission_ids
            )
        else:
            cursor.execute("""
                SELECT s.submission_id, s.raw_text FROM submissions s
                LEFT JOIN submission_themes st ON s.submission_id = st.submission_id
                WHERE st.submission_id IS NULL
            """)
            
        rows = cursor.fetchall()
        if not rows:
            return {"processed_count": 0, "theme_assignments": []}
            
        df_subs = pd.DataFrame([dict(r) for r in rows])
        texts = df_subs['raw_text'].tolist()
        
        use_fallback = True
        labels = None
        probabilities = None
        
        # We only try HDBSCAN if we have at least 3 items to group
        if len(texts) >= 3:
            try:
                from sentence_transformers import SentenceTransformer
                from sklearn.cluster import HDBSCAN
                
                model = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')
                embeddings = model.encode(texts)
                hdb = HDBSCAN(min_cluster_size=3)
                labels = hdb.fit_predict(embeddings)
                probabilities = hdb.probabilities_
                
                noise_ratio = np.sum(labels == -1) / len(labels)
                if noise_ratio <= 0.30:
                    use_fallback = False
            except Exception:
                # If dependencies are missing or fail, we automatically fall back
                use_fallback = True
                
        theme_assignments = []
        
        if use_fallback:
            theme_ids = {}
            for theme, keywords in themes_keywords.items():
                kw_summary = ", ".join(keywords[:8])
                theme_ids[theme] = get_or_create_theme(conn, theme, kw_summary)
            theme_ids["Unclassified"] = get_or_create_theme(conn, "Unclassified", "Unclassified or noisy items")
            
            for _, row in df_subs.iterrows():
                sub_id = row['submission_id']
                text_words = clean_and_tokenize(row['raw_text'])
                
                best_theme = "Unclassified"
                max_matches = 0
                for theme, keywords in themes_keywords.items():
                    matches = sum(1 for word in text_words if word in keywords)
                    if matches > max_matches:
                        max_matches = matches
                        best_theme = theme
                        
                confidence = min(1.0, float(max_matches) / 3.0) if max_matches > 0 else 0.0
                theme_id = theme_ids[best_theme]
                
                cursor.execute(
                    "INSERT OR REPLACE INTO submission_themes (submission_id, theme_id, confidence_score) VALUES (%s, %s, %s)",
                    (sub_id, theme_id, confidence)
                )
                cursor.execute("UPDATE submissions SET status = 'processed' WHERE submission_id = %s", (sub_id,))
                theme_assignments.append({"submission_id": sub_id, "theme_label": best_theme})
        else:
            # TF-IDF for dynamic theme mapping
            vectorizer = TfidfVectorizer(stop_words=stop_words, lowercase=True)
            tfidf_matrix = vectorizer.fit_transform(texts)
            feature_names = np.array(vectorizer.get_feature_names_out())
            
            unique_labels = [l for l in np.unique(labels) if l != -1]
            cluster_theme_ids = {}
            
            for c in unique_labels:
                indices = [i for i, label in enumerate(labels) if label == c]
                mean_tfidf = tfidf_matrix[indices].mean(axis=0).A1
                
                top_5_idx = mean_tfidf.argsort()[::-1][:5]
                top_5_words = feature_names[top_5_idx]
                keyword_summary = ", ".join(top_5_words)
                
                top_15_idx = mean_tfidf.argsort()[::-1][:15]
                top_15_words = feature_names[top_15_idx]
                
                best_theme = "Other"
                best_score = 0
                for theme, keywords in themes_keywords.items():
                    score = len(set(top_15_words).intersection(set(keywords)))
                    if score > best_score:
                        best_score = score
                        best_theme = theme
                        
                if best_theme == "Other":
                    best_theme = f"Other ({', '.join(top_5_words[:3])})"
                    
                theme_id = get_or_create_theme(conn, best_theme, keyword_summary)
                cluster_theme_ids[c] = theme_id
                
            unclassified_theme_id = get_or_create_theme(conn, "Unclassified", "Unclassified noise items")
            
            for i, label in enumerate(labels):
                sub_id = df_subs.loc[i, 'submission_id']
                if label == -1:
                    theme_id = unclassified_theme_id
                    confidence = 0.0
                    theme_label = "Unclassified"
                else:
                    theme_id = cluster_theme_ids[label]
                    confidence = float(probabilities[i])
                    cursor.execute("SELECT theme_label FROM themes WHERE theme_id = %s", (theme_id,))
                    theme_label = cursor.fetchone()['theme_label']
                    
                cursor.execute(
                    "INSERT OR REPLACE INTO submission_themes (submission_id, theme_id, confidence_score) VALUES (%s, %s, %s)",
                    (sub_id, theme_id, confidence)
                )
                cursor.execute("UPDATE submissions SET status = 'processed' WHERE submission_id = %s", (sub_id,))
                theme_assignments.append({"submission_id": sub_id, "theme_label": theme_label})
                
        conn.commit()
        return {"processed_count": len(theme_assignments), "theme_assignments": theme_assignments}
        
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error processing submissions: {e}")
    finally:
        conn.close()

@app.get("/rankings")
def get_rankings(
    village_id: Optional[str] = None,
    theme_id: Optional[int] = None,
    weight_demand: Optional[float] = None,
    weight_need: Optional[float] = None,
    weight_neglect: Optional[float] = None,
    weight_feasibility: Optional[float] = None
):
    try:
        # Build override weights dict
        weights = {}
        if weight_demand is not None:
            weights["demand"] = weight_demand
        if weight_need is not None:
            weights["need"] = weight_need
        if weight_neglect is not None:
            weights["neglect"] = weight_neglect
        if weight_feasibility is not None:
            weights["feasibility"] = weight_feasibility
            
        rankings = calculate_rankings(weights=weights if weights else None)
        
        # Apply filters if provided
        if village_id:
            rankings = [r for r in rankings if r['village_id'] == village_id]
        if theme_id is not None:
            rankings = [r for r in rankings if r['theme_id'] == theme_id]
            
        return {"rankings": rankings}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ranking calculation failed: {e}")

@app.post("/simulate-budget")
def simulate_budget(payload: BudgetSimulationRequest):
    try:
        # 1. Call calculate_rankings() with default weights
        rankings = calculate_rankings()
        
        # 2. Look up estimated costs from existing_works or defaults
        conn = get_db_connection()
        cursor = get_cursor(conn)
        cursor.execute("SELECT village_id, work_type, amount_lakh_inr FROM existing_works")
        works_rows = cursor.fetchall()
        conn.close()
        
        existing_costs = {}
        for row in works_rows:
            v_id = row['village_id']
            w_type = row['work_type']
            cost = row['amount_lakh_inr']
            theme_lbl = get_work_theme(w_type)
            if (v_id, theme_lbl) not in existing_costs:
                existing_costs[(v_id, theme_lbl)] = cost
                
        DEFAULT_THEME_COSTS = {
            "School Infrastructure": 25.0,
            "Water Supply": 18.0,
            "Road Connectivity": 40.0,
            "Healthcare": 22.0,
            "Employment/Skills": 15.0,
            "Electricity": 12.0,
            "Irrigation": 30.0
        }
        
        selected_works = []
        excluded_works = []
        
        total_budget = payload.budget_lakh_inr
        remaining_budget = total_budget
        cumulative_cost = 0.0
        
        theme_filter = payload.theme_filter
        village_filter = payload.village_filter
        
        for idx, item in enumerate(rankings):
            rank = idx + 1
            v_id = item['village_id']
            t_id = item['theme_id']
            t_label = item['theme_label']
            v_name = item['village_name']
            priority_score = item['priority_score']
            submission_count = item['submission_count']
            sample_complaints = item['sample_complaints']
            
            cost = existing_costs.get((v_id, t_label))
            if cost is None:
                cost = DEFAULT_THEME_COSTS.get(t_label, 10.0)
                
            work_item = {
                "rank": rank,
                "village_id": v_id,
                "village_name": v_name,
                "theme_label": t_label,
                "priority_score": priority_score,
                "estimated_cost_lakh": float(cost),
                "sample_complaints": sample_complaints,
                "submission_count": submission_count
            }
            
            # Apply filters
            is_filtered_out = False
            if theme_filter is not None:
                try:
                    target_theme_id = int(theme_filter)
                    if t_id != target_theme_id:
                        is_filtered_out = True
                except ValueError:
                    if t_label.lower() != theme_filter.lower():
                        is_filtered_out = True
                        
            if village_filter is not None:
                if v_id.lower() != village_filter.lower() and v_name.lower() != village_filter.lower():
                    is_filtered_out = True
                    
            if is_filtered_out:
                work_item["cumulative_cost_lakh"] = 0.0
                work_item["reason"] = "filtered_out"
                excluded_works.append(work_item)
                continue
                
            if cost <= remaining_budget:
                remaining_budget -= cost
                cumulative_cost += cost
                work_item["cumulative_cost_lakh"] = round(cumulative_cost, 2)
                selected_works.append(work_item)
            else:
                work_item["cumulative_cost_lakh"] = 0.0
                work_item["reason"] = "over_budget"
                excluded_works.append(work_item)
                
        total_allocated = round(cumulative_cost, 2)
        remaining = round(remaining_budget, 2)
        utilization_pct = round((total_allocated / total_budget * 100.0), 2) if total_budget > 0 else 0.0
        
        return {
            "total_budget_lakh": float(total_budget),
            "total_allocated_lakh": float(total_allocated),
            "remaining_lakh": float(remaining),
            "utilization_pct": float(utilization_pct),
            "selected_works": selected_works,
            "excluded_works": excluded_works
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Budget simulation failed: {e}")

@app.get("/villages/{village_id}")
def get_village(village_id: str):
    conn = get_db_connection()
    cursor = get_cursor(conn)
    
    try:
        # Fetch village info
        cursor.execute("SELECT * FROM villages WHERE village_id = %s", (village_id,))
        village_row = cursor.fetchone()
        if not village_row:
            raise HTTPException(status_code=404, detail="Village not found")
            
        village = dict(village_row)
        
        # Fetch nested data
        cursor.execute("SELECT * FROM submissions WHERE village_id = %s", (village_id,))
        village['submissions'] = [dict(r) for r in cursor.fetchall()]
        
        cursor.execute("SELECT * FROM existing_works WHERE village_id = %s", (village_id,))
        village['existing_works'] = [dict(r) for r in cursor.fetchall()]
        
        cursor.execute("SELECT * FROM schools WHERE village_id = %s", (village_id,))
        village['schools'] = [dict(r) for r in cursor.fetchall()]
        
        return village
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {e}")
    finally:
        conn.close()

# -------------------------------------------------------------------------
# Export Helpers (PDF and Excel)
# -------------------------------------------------------------------------
def create_xlsx_export(rankings):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Priority Rankings"
    ws.views.sheetView[0].showGridLines = True
    
    headers = [
        "Rank", "Village Name", "Theme", "Priority Score", 
        "Demand Score", "Need Score", "Neglect Score", "Feasibility Score", 
        "Submission Count"
    ]
    ws.append(headers)
    
    # Styled Header
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    for i, r in enumerate(rankings):
        row_data = [
            i + 1,
            r['village_name'],
            r['theme_label'],
            r['priority_score'],
            r['demand_score'],
            r['need_score'],
            r['neglect_score'],
            r['feasibility_score'],
            r['submission_count']
        ]
        ws.append(row_data)
        
        row_idx = i + 2
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx in [1, 4, 5, 6, 7, 8, 9]:
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def create_pdf_export(rankings):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#1F497D'),
        spaceAfter=12
    )
    body_style = ParagraphStyle(
        'BodyStyle',
        parent=styles['Normal'],
        fontSize=10,
        leading=13,
        spaceAfter=6
    )
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.white,
        fontName='Helvetica-Bold'
    )
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontSize=8,
        leading=10
    )

    elements = []
    
    elements.append(Paragraph("People's Priorities AI - Prioritization Rankings", title_style))
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}. Summarizing ranked community needs across villages.", body_style))
    elements.append(Spacer(1, 10))
    
    table_data = [[
        Paragraph("Rank", table_header_style),
        Paragraph("Village", table_header_style),
        Paragraph("Theme", table_header_style),
        Paragraph("Priority", table_header_style),
        Paragraph("Demand", table_header_style),
        Paragraph("Need", table_header_style),
        Paragraph("Neglect", table_header_style),
        Paragraph("Feasibility", table_header_style),
        Paragraph("Count", table_header_style)
    ]]
    
    for i, r in enumerate(rankings):
        table_data.append([
            Paragraph(str(i + 1), table_cell_style),
            Paragraph(r['village_name'], table_cell_style),
            Paragraph(r['theme_label'], table_cell_style),
            Paragraph(f"{r['priority_score']:.2f}", table_cell_style),
            Paragraph(f"{r['demand_score']:.1f}", table_cell_style),
            Paragraph(f"{r['need_score']:.1f}", table_cell_style),
            Paragraph(f"{r['neglect_score']:.1f}", table_cell_style),
            Paragraph(f"{r['feasibility_score']:.1f}", table_cell_style),
            Paragraph(str(r['submission_count']), table_cell_style)
        ])
        
    col_widths = [30, 80, 110, 50, 50, 50, 50, 60, 60]
    
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F497D')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D3D3D3')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F5F8')])
    ]))
    
    elements.append(t)
    doc.build(elements)
    
    buffer.seek(0)
    return buffer

@app.get("/export")
def export_rankings(
    format: str = Query(..., pattern="^(pdf|xlsx)$", description="Format to export: pdf or xlsx"),
    village_id: Optional[str] = None
):
    try:
        rankings = calculate_rankings()
        if village_id:
            rankings = [r for r in rankings if r['village_id'] == village_id]
            
        if format == "xlsx":
            buffer = create_xlsx_export(rankings)
            filename = f"priorities_report_{village_id or 'all'}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            buffer = create_pdf_export(rankings)
            filename = f"priorities_report_{village_id or 'all'}.pdf"
            media_type = "application/pdf"
            
        return StreamingResponse(
            buffer,
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {e}")

# -------------------------------------------------------------------------
# Budget Simulation Export Helpers & Endpoint
# -------------------------------------------------------------------------
def create_budget_xlsx_export(sim_result):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    
    # Sheet 1 "Sanction List"
    ws1 = wb.active
    ws1.title = "Sanction List"
    headers1 = ["Rank", "Village", "Development Work", "Estimated Cost (₹ Lakhs)", "Priority Score", "Citizen Submissions"]
    ws1.append(headers1)
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'), top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3'))
    
    for col_idx in range(1, len(headers1) + 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    alt_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    
    for i, work in enumerate(sim_result["selected_works"]):
        row_data = [
            work['rank'],
            work['village_name'],
            work['theme_label'],
            work['estimated_cost_lakh'],
            work['priority_score'],
            work['submission_count']
        ]
        ws1.append(row_data)
        row_idx = i + 2
        for col_idx in range(1, len(headers1) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if i % 2 != 0:
                cell.fill = alt_fill
                
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Sheet 2 "Summary"
    ws2 = wb.create_sheet(title="Summary")
    summary_data = [
        ("Total Budget (₹ Lakhs)", sim_result["total_budget_lakh"]),
        ("Total Allocated (₹ Lakhs)", sim_result["total_allocated_lakh"]),
        ("Remaining Budget (₹ Lakhs)", sim_result["remaining_lakh"]),
        ("Utilization %", sim_result["utilization_pct"]),
        ("Number of Works Selected", len(sim_result["selected_works"])),
        ("Date Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    ]
    for k, v in summary_data:
        ws2.append([k, v])
    for row in ws2.iter_rows():
        row[0].font = Font(bold=True)
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 20
        
    # Sheet 3 "Excluded Works"
    ws3 = wb.create_sheet(title="Excluded Works")
    headers3 = ["Rank", "Village", "Development Work", "Estimated Cost (₹ Lakhs)", "Priority Score", "Reason"]
    ws3.append(headers3)
    for col_idx in range(1, len(headers3) + 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        
    for i, work in enumerate(sim_result["excluded_works"]):
        row_data = [
            work['rank'],
            work['village_name'],
            work['theme_label'],
            work['estimated_cost_lakh'],
            work['priority_score'],
            work.get('reason', '')
        ]
        ws3.append(row_data)
    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws3.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def create_budget_pdf_export(sim_result):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    buffer = BytesIO()
    
    def add_page_border(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor('#1F497D'))
        canvas.setLineWidth(2)
        canvas.rect(18, 18, letter[0] - 36, letter[1] - 36)
        canvas.restoreState()
        
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=16,
        leading=20,
        alignment=1, # Center
        textColor=colors.HexColor('#1F497D'),
        spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Heading2'],
        fontSize=10,
        alignment=1, # Center
        textColor=colors.HexColor('#555555'),
        spaceAfter=12
    )
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        spaceBefore=20,
        alignment=0
    )
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.white,
        fontName='Helvetica-Bold',
        alignment=1
    )
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontSize=8,
        alignment=1
    )
    
    elements = []
    
    elements.append(Paragraph("MPLADS Development Priority Sanction List", title_style))
    elements.append(Paragraph(f"District: Simulated District | Total Budget: ₹ {sim_result['total_budget_lakh']} Lakhs | Date: {datetime.now().strftime('%Y-%m-%d')}", subtitle_style))
    elements.append(Spacer(1, 10))
    
    table_data = [[
        Paragraph("Rank", table_header_style),
        Paragraph("Village", table_header_style),
        Paragraph("Development Work", table_header_style),
        Paragraph("Estimated Cost (₹ Lakhs)", table_header_style),
        Paragraph("Priority Score", table_header_style),
        Paragraph("Citizen Submissions", table_header_style)
    ]]
    
    for r in sim_result['selected_works']:
        table_data.append([
            Paragraph(str(r['rank']), table_cell_style),
            Paragraph(r['village_name'], table_cell_style),
            Paragraph(r['theme_label'], table_cell_style),
            Paragraph(str(r['estimated_cost_lakh']), table_cell_style),
            Paragraph(f"{r['priority_score']:.2f}", table_cell_style),
            Paragraph(str(r['submission_count']), table_cell_style)
        ])
        
    col_widths = [40, 100, 150, 90, 70, 90]
    
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F497D')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D3D3D3')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F5F8')])
    ]))
    
    elements.append(t)
    
    footer_text = f"<b>Total Allocated:</b> ₹ {sim_result['total_allocated_lakh']} Lakhs<br/>"
    footer_text += f"<b>Remaining Budget:</b> ₹ {sim_result['remaining_lakh']} Lakhs<br/><br/>"
    footer_text += "<i>Generated by People's Priorities AI</i>"
    elements.append(Paragraph(footer_text, footer_style))
    
    doc.build(elements, onFirstPage=add_page_border, onLaterPages=add_page_border)
    
    buffer.seek(0)
    return buffer

@app.post("/simulate-budget/export")
def simulate_budget_export_endpoint(payload: BudgetSimulationExportRequest):
    try:
        base_payload = BudgetSimulationRequest(
            budget_lakh_inr=payload.budget_lakh_inr,
            theme_filter=payload.theme_filter,
            village_filter=payload.village_filter
        )
        sim_result = simulate_budget(base_payload)
        
        if payload.format == "xlsx":
            buffer = create_budget_xlsx_export(sim_result)
            filename = "sanction_list.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            buffer = create_budget_pdf_export(sim_result)
            filename = "sanction_list.pdf"
            media_type = "application/pdf"
            
        return StreamingResponse(
            buffer,
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Budget Export failed: {e}")
